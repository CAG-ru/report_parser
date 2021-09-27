#!/usr/bin/env python3
import os
import tempfile
from pathlib import Path

import mammoth
from openpyxl import load_workbook
from pandas import DataFrame
from xls2xlsx import XLS2XLSX
from xlsx2html import xlsx2html

from src.documents_converter import convert
from src.html_utils import get_texts_and_tables
from src.parsing_tree import ParsingTree
from src.text_class import Text


class Report:
    def __init__(self, fname):
        self.fname = fname
        self.name = Path(fname).name
        self.ext = Path(fname).suffix.strip('.')
        self.log = []
        self.contents = []
        self.html = []
        if self.ext not in ['html', 'doc', 'docx', 'xls', 'xlsx']:
            self.log.append(f'Расширение {self.ext} не поддерживается')
        self.load()

    def load(self):
        with open(self.fname, 'rb') as f:
            if self.ext == 'docx':
                converted = mammoth.convert(f)
                self.log += converted.messages
                self.html.append(converted.value)

            elif self.ext == 'xlsx':
                for _ in load_workbook(f).sheetnames:
                    converted = xlsx2html(f, sheet=_)
                    converted.seek(0)
                    self.html.append(converted.read())

            elif self.ext == 'xls':
                with tempfile.TemporaryFile() as f:
                    x2x_file = XLS2XLSX(self.fname).to_xlsx(f)
                    for _ in load_workbook(f).sheetnames:
                        converted = xlsx2html(x2x_file, sheet=_)
                        converted.seek(0)
                        self.html.append(converted.read())
            elif self.ext == 'doc':
                docx = convert.convert(self.fname, 'docx')

                temp_file = f'.{self.name}.docx'
                with open(temp_file, "wb") as fp:
                    fp.write(docx)
                with open(temp_file, 'rb') as fp:
                    converted = mammoth.convert(fp)
                    self.log += converted.messages
                    self.html.append(converted.value)
                os.remove(temp_file)

            else:
                self.html.append(f.read())

    def texts(self):
        return (x for x in self.contents if type(x) == Text)

    def tables(self):
        return (x for x in self.contents if type(x) == DataFrame)

    def parse(self, new_method=False):
        if not self.html:
            return
        for _ in self.html:
            tree = ParsingTree(_)
            html_elems = tree.elements
            self.contents.extend(get_texts_and_tables(html_elems, new_method))

    def __str__(self):
        to_str = ''
        for elem in self.contents:
            to_str += str(type(elem)) + '\n'
            to_str += str(elem) + '\n'
        return to_str

    def find_table_by_name(self, name):
        add_next_table = False
        find_text = True
        result = []
        for elem in self.contents:
            if type(elem) == Text:
                if find_text and elem.contains(name):
                    add_next_table = True
                    find_text = False
                else:
                    add_next_table = False
            else:
                if add_next_table:
                    result.append(elem)
        return result


if __name__ == '__main__':
    examples = ['doc_report.doc']
    for fn in examples:
        print(fn)
        path = 'examples/' + fn
        report = Report(path)
        report.parse(new_method=True)
        print('log:\n', report.log)
        for text in report.texts():
            print(text)
        for table in report.tables():
            print(table)
